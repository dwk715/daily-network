#!/usr/bin/python
# -*- coding: utf-8 -*-
# Date: 2018/11/5
# Autor :  zlw dwk zly 

import os
import openpyxl
import time

'''
此模块用于将数据保存在excel文件中
读取excel/template.xlsx文件并另存为：
excel/易盛上海分公司日常巡检表_{日期}.xlsx
'''


'''
打开表格文件
return dict 包含需要保存的文件名，模板表格对象、表格最大行
'''

def open_xlsx():
    filename = 'excel/易盛上海分公司日常巡检表_' + time.strftime('%Y-%m-%d',
                                               time.localtime()) + '.xlsx'
    if (os.path.exists(filename)):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.load_workbook('excel/template.xlsx')
    ws = wb.active
    max = ws.max_row
    return ({"filename": filename, "wb": wb, "max_raw": max})


'''
保存cpu、内存信息
device：str 设备名
cpu_mem_result：dict eg:{"cpu":40,"mem":60}
'''


def cpu_mem(device, cpu_mem_result):
    wb_info = open_xlsx()
    wb = wb_info["wb"]
    ws = wb.active
    for work_row in range(5, wb_info['max_raw']):
        if (ws.cell(row=work_row, column=1).value == device):
            ws.cell(row=work_row, column=5).value = str(cpu_mem_result['cpu'])
            ws.cell(row=work_row, column=7).value = str(cpu_mem_result['mem'])
    wb.save(wb_info["filename"])


'''
保存可用端口数
device：str 设备名
interface_result：int/str 可用端口数
'''


def interface(device, interface_result):
    wb_info = open_xlsx()
    wb = wb_info["wb"]
    ws = wb.active
    for work_row in range(5, wb_info['max_raw']):
        if (ws.cell(row=work_row, column=1).value == device):
            ws.cell(row=work_row, column=3).value = str(interface_result)
    wb.save(wb_info["filename"])


'''
根据当前时间保存流量结果
device：str 设备名
flow_result：list eg:[{'in': '17.0', 'out': '51.0'},{'in': '8.0', 'out': '32.0'}]
'''


def flow(device, flow_result):
    hour=time.strftime('%H',time.localtime())
    if (hour==10):
        col=7
    else if (hour==22):
        col=8
    wb_info = open_xlsx()
    wb = wb_info["wb"]
    ws = wb.active
    write_rows = []
    for work_row in range(5, wb_info['max_raw']):
        if (ws.cell(row=work_row, column=4).value == device):
            write_rows.append(work_row)
    for i, v in enumerate(flow_result):
        ws.cell(row=write_rows[i], column=col).value = str(v['in'])
        ws.cell(row=write_rows[i], column=(col + 2)).value = str(v['out'])
    wb.save(wb_info["filename"])


'''
保存ping结果
device：str 设备名
ping_result：list 处理后的ping数据 eg [{"loss":0,"time":6},{"loss":1,"time":7}]
'''


def ping(device, ping_result):
    wb_info = open_xlsx()
    wb = wb_info["wb"]
    ws = wb.active
    write_rows = []
    for work_row in range(5, 32):
        #print (work_row,ws.cell(row=work_row, column=4).value)
        if (ws.cell(row=work_row, column=4).value == device):
            write_rows.append(work_row)
    for i, v in enumerate(ping_result):
        if (v['loss'] > 0):
            ws.cell(
                row=write_rows[i],
                column=5).fill = openpyxl.styles.PatternFill(
                    "solid", fgColor="FFC125")
        ws.cell(row=write_rows[i], column=5).value = str(v['loss'])
        ws.cell(row=write_rows[i], column=6).value = str(v['avg'])
    wb.save(wb_info["filename"])