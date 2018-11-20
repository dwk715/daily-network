#!/usr/bin/python
# -*- coding: utf-8 -*-
# Date: 2018/11/5
# Author :  zlw dwk zly

import os
import openpyxl
import time
from .log import log_instance
from .slack_bot import dn_say
import traceback
from pymongo import MongoClient

'''
此模块用于将数据保存在excel文件中
读取excel/template.xlsx文件并另存为：
excel/易盛上海分公司日常巡检表_{日期}.xlsx
'''
'''
打开表格文件
return dict 包含需要保存的文件名，模板表格对象、表格最大行
'''

line_name_convert = {
    "上海电信": "SH_CNTC",
    "上海联通": "SH_CNUC",
    "香港电信": "HK_CNTC",
    "北京联通": "BJ_CNUC",
    "电讯盈科": "HK_PCCW",
    "沪港联通": "HK_CNUC",
    "深圳移动": "SZ_CMCC",
    "北京数讯": "BJ_SX",
    "上海电信（主）": "SH_CNTC_MASTER",
    "上海移动（主）": "SH_CMCC_MASTER",
    "上海移动（备）": "SH_CMCC_BACKUP",
    "上海电信（备）": "SH_CNTC_BACKUP",
}

try:
    Client = MongoClient('mongodb://127.0.0.1:27017/')
    db = Client['daily_network']
    collection_line = db['line']
    collection_device = db['device']
except Exception as e:
    print(e)
    dn_say(traceback.format_exc())


def open_excel():
    filename = 'excel/易盛上海分公司日常巡检表_' + time.strftime(
        '%Y-%m-%d', time.localtime()) + '.xlsx'
    open_file = filename if os.path.exists(filename) else 'excel/template.xlsx'
    wb_info = {}
    try:
        wb = openpyxl.load_workbook(open_file)
        ws = wb.active
        if open_file == 'excel/template.xlsx':
            ws.cell(
                row=2, column=1).value = time.strftime('日期: %Y 年 %m 月 %d 日',
                                                       time.localtime())
        max = ws.max_row
        wb_info.update({"filename": filename, "wb": wb, "ws": ws, "max_row": max})
    except IOError as e:
        log_instance.critical("open file error!", e)
        dn_say(traceback.format_exc())
    return wb_info


def read_db_to_write_excel():
    wb_info = open_excel()
    ws = wb_info['ws']
    max_row = wb_info['max_row']
    for row in range(5, max_row):
        device_name = ws.cell(row=row, column=1).value
        line_name = ws.cell(row=row, column=2).value
        # device部分写入
        if collection_device.count_documents({'name': device_name}) == 1:
            device_info = collection_device.find_one({'name': device_name})
            if device_info['interface']:
                # 剩余接口写入
                ws.cell(row=row, column=3).value = str(device_info['interface'][-1]['available'])
                # cpu使用率写入
                ws.cell(row=row, column=5).value = str(device_info['cpu'][-1]['use'])
                # memory剩余写入
                ws.cell(row=row, column=7).value = str(device_info['memory'][-1]['remain'])
            else:
                # cpu使用率写入
                ws.cell(row=row, column=5).value = str(device_info['cpu'][-1]['use'])
                # memory剩余写入
                ws.cell(row=row, column=7).value = str(device_info['memory'][-1]['remain'])
        # line部分写入
        if line_name in line_name_convert.keys():
            if collection_line.count_documents({'name': line_name_convert[line_name]}) == 1:
                line_info = collection_line.find_one({'name': line_name_convert[line_name]})
                # 丢包率写入：
                ws.cell(row=row, column=5).value = str(line_info['loss'][-1]['value'])
                # 平均延时写入:
                ws.cell(row=row, column=6).value = str(line_info['delay'][-1]['value'])
                # 早间流量上行写入
                ws.cell(row=row, column=7).value = str(line_info['flow_out_am'][-1]['value'])
                # 晚间流量上行写入
                ws.cell(row=row, column=8).value = str(line_info['flow_out_pm'][-1]['value'])
                # 早间流量下行写入
                ws.cell(row=row, column=9).value = str(line_info['flow_in_am'][-1]['value'])
                # 晚间流量下行写入
                ws.cell(row=row, column=10).value = str(line_info['flow_in_pm'][-1]['value'])
    wb_info["wb"].save(wb_info["filename"])
