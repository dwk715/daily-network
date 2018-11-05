#!/usr/bin/python
# -*- coding: utf-8 -*-
# Date: 2018/10/31

from netmiko import ConnectHandler
import datetime
import time
import os
import csv
import logging
import yaml
import re
import openpyxl
from log import logmode
from apscheduler.schedulers.blocking import BlockingScheduler

#创建一个log实例
logger = logmode('daily-work').getlog()
log_ap = logmode('apscheduler').getlog()


#login and resolve result
def connect_devices(device_info, commands):
    ip = device_info[0]
    protocol = device_info[1]
    device_type = device_info[2]
    username = device_info[3]
    password = device_info[4]
    secret = device_info[5]
    port = 22 if protocol == 'ssh' else 23

    device_type = device_type + '_telnet' if protocol == 'telnet' else device_type

    cisco_device = {
        'device_type': device_type,
        'ip': ip,
        'username': username,
        'password': password,
        'port': port,  # optional, defaults to 22
        'secret': secret,  # optional, defaults to ''
        'verbose': False,  # optional, defaults to False
    }

    net_connect = ConnectHandler(**cisco_device)
    net_connect.enable()

    result = []
    for cmd in commands:
        result.append(net_connect.send_command(cmd))

    net_connect.disconnect()
    return result


def parse_ping(result):
    ping_result_write = []
    for row in result:
        if row == '':
            pass
        else:
            matchObj = re.search(
                r'.* rate is (.*?) .*max = (.*?)/(.*?)/(.*?).*', row,
                re.M | re.I)
            percent = 100 - int(matchObj.group(1))
            avg = matchObj.group(3)

            ping_result = {
                'loss': percent,
                'avg': avg,
            }
            ping_result_write.append(ping_result)

    return ping_result_write


def parse_cpu_mem(cpu_mem_raw):
    result = {}
    cpu_reg = re.compile(r'.*?utilization.*?:\s*?(?P<cpu>\d.*?)%.*',
                         re.S | re.M)
    switch_mem_reg = re.compile(
        r'.*?Pool Total:\s*?(?P<total>\d+)\s*.*?Free:\s*?(?P<free>\d+)\s*')
    router_mem_reg = re.compile(r'.*?\((?P<used_rate>\d+)%\).*')
    firewall_mem_reg = re.compile(
        r'.*?Free memory:.*?\((?P<free_rate>\d+)%\).*')
    cpu_regMatch = cpu_reg.match(cpu_mem_raw[1])
    mem_regMatch = switch_mem_reg.match(
        cpu_mem_raw[2]) or firewall_mem_reg.match(
            cpu_mem_raw[2]) or router_mem_reg.match(cpu_mem_raw[2])
    result['cpu'] = cpu_regMatch.groupdict()['cpu']
    #print(mem_regMatch.groupdict())
    if 'used_rate' in mem_regMatch.groupdict():
        result['mem'] = 100 - (float(mem_regMatch.groupdict()['used_rate']))
    if 'free_rate' in mem_regMatch.groupdict():
        result['mem'] = float(mem_regMatch.groupdict()['free_rate'])
    if 'total' in mem_regMatch.groupdict():
        #print ('free',mem_regMatch.groupdict()['free'],'tatal',mem_regMatch.groupdict()['total'])
        result['mem'] = format(
            float(mem_regMatch.groupdict()['free']) * 100 / float(
                mem_regMatch.groupdict()['total']), "0.1f")
    #print (result['cpu'],result['mem'])
    #save_cpu_mem(device,result)
    return result


'''
解析可用接口数
interface_raw：list [1]为接口原始信息
return int 可用接口数
'''


def parse_interface(interface_raw):
    #save_interface(device,len(re.findall("Ethernet",interface_raw[1])))
    return len(re.findall("Ethernet", interface_raw[1]))


'''
解析流量
flow_raw：list 流量原始信息
return list 处理后的流量数据 eg:[{'in': '17.0', 'out': '51.0'},{'in': '8.0', 'out': '32.0'}]
'''


def parse_flow(flow_raw):
    results = []
    flow_reg=re.compile(r'''.*?input.*?\s+(?P<in>\d+?)\s+b.*?/sec.*
  .*?output.*?\s+(?P<out>\d+?)\s+b.*?/sec.*''', re.S|re.M)
    for flow_result in flow_raw:
        if not flow_result.strip():
            continue
        result = {}
        regMatch = flow_reg.match(flow_result)
        linebits = regMatch.groupdict()
        for k, v in linebits.items():
            if 'bytes' in flow_result:
                v = format(
                    float(re.findall("\d+", v)[0]) * 8 / (1024 * 1024), "0.1f")
                result[k] = v
            if 'bit' in flow_result:
                v = format(
                    float(re.findall("\d+", v)[0]) / (1024 * 1024), "0.1f")
                result[k] = v
        results.append(result)
    #save_flow(device,results)
    return results


#打开表格文件，返回需要打开的表格对象、表格最大行、需要保存的文件名
def open_xlsx():
    filename = '易盛上海分公司日常巡检表_' + time.strftime('%Y-%m-%d',
                                               time.localtime()) + '.xlsx'
    if (os.path.exists(filename)):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.load_workbook('易盛上海分公司日常巡检表V3.3.xlsx')
    ws = wb.active
    max = ws.max_row
    return ({"filename": filename, "wb": wb, "max_raw": max})


'''
保存cpu、内存信息
device：str 设备名
cpu_mem_result：dict eg:{"cpu":40,"mem":60}
'''


def save_cpu_mem(device, cpu_mem_result):
    wb_info = open_xlsx()
    wb = wb_info["wb"]
    ws = wb.active
    for work_row in range(5, wb_info['max_raw']):
        if (ws.cell(row=work_row, column=1).value == device):
            ws.cell(row=work_row, column=5).value = str(cpu_mem_result['cpu'])
            ws.cell(row=work_row, column=7).value = str(cpu_mem_result['mem'])
    wb.save(wb_info["filename"])


'''
保存可用端口数
device：str 设备名
interface_result：int/str 可用端口数
'''


def save_interface(device, interface_result):
    wb_info = open_xlsx()
    wb = wb_info["wb"]
    ws = wb.active
    for work_row in range(5, wb_info['max_raw']):
        if (ws.cell(row=work_row, column=1).value == device):
            ws.cell(row=work_row, column=3).value = str(interface_result)
    wb.save(wb_info["filename"])


'''
保存流量结果
device：str 设备名
flow_result：list eg:[{'in': '17.0', 'out': '51.0'},{'in': '8.0', 'out': '32.0'}]
'''


def save_flow(device, flow_result):
    wb_info = open_xlsx()
    wb = wb_info["wb"]
    ws = wb.active
    write_rows = []
    for work_row in range(5, wb_info['max_raw']):
        if (ws.cell(row=work_row, column=4).value == device):
            write_rows.append(work_row)
    for i, v in enumerate(flow_result):
        col = 8 if (ws.cell(row=write_rows[i], column=7).value) else 7
        ws.cell(row=write_rows[i], column=col).value = str(v['in'])
        ws.cell(row=write_rows[i], column=(col + 2)).value = str(v['out'])
    wb.save(wb_info["filename"])


'''
保存ping结果
device：str 设备名
ping_result：list 处理后的ping数据 eg [{"loss":0,"time":6},{"loss":1,"time":7}]
'''


def save_ping(device, ping_result):
    wb_info = open_xlsx()
    wb = wb_info["wb"]
    ws = wb.active
    write_rows = []
    for work_row in range(5, 32):
        #print (work_row,ws.cell(row=work_row, column=4).value)
        if (ws.cell(row=work_row, column=4).value == device):
            write_rows.append(work_row)
    for i, v in enumerate(ping_result):
        if (v['loss'] > 0):
            ws.cell(
                row=write_rows[i],
                column=5).fill = openpyxl.styles.PatternFill(
                    "solid", fgColor="FFC125")
        ws.cell(row=write_rows[i], column=5).value = str(v['loss'])
        ws.cell(row=write_rows[i], column=6).value = str(v['avg'])
    wb.save(wb_info["filename"])


#read the csv
def read_csv(ip, commands):
    result = []
    with open('config.csv', mode='r') as f:
        # 逐行读入csv
        f_csv = csv.reader(f)
        for row in f_csv:
            if ip == row[0]:
                result = connect_devices(row, commands)
            else:
                pass
    return result



def read_yml(tables):
    path = "commands/" + tables
    files = os.listdir(path)
    for file in files:
        if not os.path.isdir(file):
            f = open(path + "/" + file)
            y = yaml.load(f)
            tmp = read_csv(y['ip'], y['commands'])
            if tables == 'ping':
                save_ping(y['device_name'], parse_ping(tmp))

            if tables == 'cpu_memory':
                save_cpu_mem(y['device_name'], parse_cpu_mem(tmp))

            if tables == 'flow':
                # print(y['device_name'], tmp)
                save_flow(y['device_name'], parse_flow(tmp))

            if tables == 'interface':
                save_interface(y['device_name'], parse_interface(tmp))

        else:
            pass


def job_ping():
    read_yml('ping')


def job_flow():
    read_yml('flow')
  

def job_cup_memory():
    read_yml('cpu_memory')


def job_interface():
    read_yml('interface')
  

def main():
    scheduler = BlockingScheduler()
    scheduler.add_job(job_interface, 'cron', day_of_week='0-6', hour=8, minute=00)
    scheduler.add_job(job_cup_memory, 'cron', day_of_week='0-6', hour=10, minute=30)
    scheduler.add_job(job_flow, 'cron', day_of_week='0-6', hour=10, minute=30)
    scheduler.add_job(job_flow, 'cron', day_of_week='0-6', hour=22, minute=20)
    scheduler.add_job(job_ping, 'cron', day_of_week='0-6', hour=8, minute=00)
    # read_yml('interface')
    scheduler.start()


if __name__ == '__main__':
    main()
